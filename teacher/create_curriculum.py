import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Python課綱"

GREEN_DARK  = "1E5631"
GREEN_MED   = "4CAF50"
GREEN_LIGHT = "C8E6C9"
BROWN       = "795548"
BROWN_LIGHT = "D7CCC8"
GOLD        = "FFC107"
GOLD_LIGHT  = "FFF9C4"
STONE_LIGHT = "F5F5F5"
WHITE       = "FFFFFF"
ORANGE_LT   = "FFE0B2"
PURPLE_LT   = "E1BEE7"

def fill(hex_c): return PatternFill("solid", fgColor=hex_c)
def border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)
def font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, name="Calibri", size=size)

# ── Title ──────────────────────────────────────────────────
ws.merge_cells("A1:F1")
t = ws["A1"]
t.value     = "⛏️  Python 初學者課綱（12堂課 × 1小時）  Minecraft Edition  ⛏️"
t.font      = Font(bold=True, size=15, color=WHITE, name="Calibri")
t.fill      = fill(GREEN_DARK)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 34

# ── Headers ────────────────────────────────────────────────
headers    = ["堂次", "時間", "課程主題", "課程內容（重點）", "課堂練習", "作業"]
col_widths = [8, 10, 24, 50, 36, 38]
for c, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=2, column=c, value=h)
    cell.font      = font(bold=True, color=WHITE, size=11)
    cell.fill      = fill(GREEN_MED)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = border()
    ws.column_dimensions[get_column_letter(c)].width = w
ws.row_dimensions[2].height = 26

# ── Lesson data ────────────────────────────────────────────
sessions = [
    {
        "no"      : "第1堂",
        "time"    : "1 小時",
        "topic"   : "🐍 認識 Python & print()",
        "content" : "1. Python 是什麼？能做什麼？\n2. 安裝 Python / VS Code\n3. print() 印出文字\n4. 字串的概念（引號）\n5. # 註解",
        "practice": "• 印出角色名稱\n• 用 print() 做角色卡\n• 試試忘記引號會怎樣",
        "hw"      : "「Minecraft 自我介紹」\n用 print() 印出角色卡\n包含名稱、愛好、武器、目標",
        "fill"    : BROWN_LIGHT,
    },
    {
        "no"      : "第2堂",
        "time"    : "1 小時",
        "topic"   : "🟫 變數 & 資料型別",
        "content" : "1. 變數是什麼（背包格子比喻）\n2. int / float / str / bool\n3. type() 查看型別\n4. 更新變數值\n5. +=  -=  *=  縮寫運算",
        "practice": "• 建立角色 hp/name/level\n• 模擬戰鬥（hp 扣血）\n• 猜 type() 的結果",
        "hw"      : "「角色卡 v1」\n建立 6 個不同型別的變數\n模擬角色受傷、回血、升級",
        "fill"    : BROWN_LIGHT,
    },
    {
        "no"      : "第3堂",
        "time"    : "1 小時",
        "topic"   : "💬 字串操作",
        "content" : "1. 字串串接 +\n2. 字串重複 *\n3. len() 長度\n4. upper() / lower() / replace()\n5. strip() 去空白",
        "practice": "• 做 Minecraft 告示牌分隔線\n• 字串方法練習\n• 合成武器名稱字串",
        "hw"      : "「附魔武器產生器」\n建立武器+附魔變數\n印出格式化武器說明卡",
        "fill"    : STONE_LIGHT,
    },
    {
        "no"      : "第4堂",
        "time"    : "1 小時",
        "topic"   : "⌨️  input() & f-string & 型別轉換",
        "content" : "1. input() 接收輸入\n2. input() 永遠回傳 str\n3. int() / float() / str() 轉換\n4. f-string 格式化輸出\n5. {} 裡可放運算式",
        "practice": "• 互動角色產生器\n• 計算礦物總價值\n• 兩數相加程式",
        "hw"      : "「角色卡 v2（互動版）」\n用 input 讓使用者輸入\n計算攻擊力、印出角色卡",
        "fill"    : STONE_LIGHT,
    },
    {
        "no"      : "第5堂",
        "time"    : "1 小時",
        "topic"   : "⚔️  條件判斷 if / elif / else",
        "content" : "1. if 語法與縮排\n2. if ... else\n3. if ... elif ... else\n4. 比較運算子 > < >= <= == !=\n5. = 和 == 的差異",
        "practice": "• 血量狀態判斷\n• 合成台材料檢查\n• Minecraft 時間判斷",
        "hw"      : "「Minecraft 生存顧問」\n輸入血量/食物/武器\n給出生存建議",
        "fill"    : GOLD_LIGHT,
    },
    {
        "no"      : "第6堂",
        "time"    : "1 小時",
        "topic"   : "🔮 邏輯運算子 & 巢狀 if",
        "content" : "1. and（兩個都要成立）\n2. or（一個成立就好）\n3. not（反轉條件）\n4. 組合多個條件\n5. 巢狀 if（if 裡面的 if）",
        "practice": "• Boss 戰前裝備判斷\n• 礦物鑑定師\n• 多條件生存判斷",
        "hw"      : "「Minecraft 合成台程式」\n輸入材料數量\n判斷可合成哪些武器",
        "fill"    : GOLD_LIGHT,
    },
    {
        "no"      : "第7堂",
        "time"    : "1 小時",
        "topic"   : "🔄  for 迴圈",
        "content" : "1. 為什麼需要迴圈\n2. for i in range() 語法\n3. range(start, stop, step)\n4. break 跳出 / continue 跳過\n5. for + if 組合\n6. % 取餘數運算子（奇偶判斷）",
        "practice": "• 挖礦計數器\n• 礦層分析\n• 加總礦物數量",
        "hw"      : "「自動礦物統計機」\n for 迴圈跑各層\n統計鑽石/鐵礦層數",
        "fill"    : GREEN_LIGHT,
    },
    {
        "no"      : "第8堂",
        "time"    : "1 小時",
        "topic"   : "🔁  while 迴圈",
        "content" : "1. for vs while 的差異\n2. while 語法與條件\n3. 無限迴圈陷阱（忘了更新）\n4. while True + break 模式\n5. 互動式選單",
        "practice": "• 猜數字（鑽石Y層）\n• 戰鬥模擬器\n• 背包選單雛型",
        "hw"      : "「Minecraft 生存模擬器」\n選單：挖礦/狩獵/吃東西\n金幣/血量/食物值管理",
        "fill"    : PURPLE_LT,
    },
    {
        "no"      : "第9堂",
        "time"    : "1 小時",
        "topic"   : "🎒  串列 List",
        "content" : "1. List 是什麼（Minecraft 背包）\n2. 索引（從 0 開始）\n3. append / remove / pop\n4. len() / in 運算子\n5. 切片 [start:end]",
        "practice": "• 背包物品管理\n• 計算背包裡的鑽石數\n• 排序背包",
        "hw"      : "「完整背包管理系統」\n新增/刪除/查詢/排序\n背包上限 9 格",
        "fill"    : ORANGE_LT,
    },
    {
        "no"      : "第10堂",
        "time"    : "1 小時",
        "topic"   : "🔧  函式 Function",
        "content" : "1. 為什麼要用函式\n2. def 定義語法\n3. 參數 (parameter)\n4. return 回傳值\n5. 預設參數值",
        "practice": "• 傷害計算函式\n• 角色卡顯示函式\n• 合成判斷函式",
        "hw"      : "「Minecraft RPG 工具包」\n寫 5 個函式（狀態/傷害/\n合成/升級/is_alive）",
        "fill"    : STONE_LIGHT,
    },
    {
        "no"      : "第11堂",
        "time"    : "1 小時",
        "topic"   : "🏹  綜合練習 for+if+while+list+function",
        "content" : "1. 快問快答複習\n2. 礦物統計機（全語法混合）\n3. Debug 練習（找出 3 個 bug）\n4. Minecraft RPG 戰鬥系統\n5. 設計技巧：拆解問題",
        "practice": "• 猜輸出快問快答\n• 找 bug 練習\n• 一起寫戰鬥系統",
        "hw"      : "「Minecraft 地下城探索」\n3 個房間、怪物、寶箱\n戰鬥/逃跑選擇",
        "fill"    : GOLD_LIGHT,
    },
    {
        "no"      : "第12堂",
        "time"    : "1 小時",
        "topic"   : "🏆  期末專案：Minecraft 文字冒險",
        "content" : "1. 展示第11堂作業\n2. 說明期末專案架構\n3. 一起閱讀和理解範本\n4. 學生嘗試修改/擴充\n5. 課程總結 & 下一步",
        "practice": "• 展示作業\n• 讀懂期末範本程式碼\n• 嘗試加入新房間或功能",
        "hw"      : "【期末專案】完整版文字冒險\n至少 3 個場景、怪物戰鬥\n背包系統、勝利/失敗結局",
        "fill"    : GREEN_LIGHT,
    },
]

row_fills = [
    BROWN_LIGHT, BROWN_LIGHT, STONE_LIGHT, STONE_LIGHT,
    GOLD_LIGHT,  GOLD_LIGHT,  GREEN_LIGHT, PURPLE_LT,
    ORANGE_LT,   STONE_LIGHT, GOLD_LIGHT,  GREEN_LIGHT,
]

for i, s in enumerate(sessions):
    r = i + 3
    ws.row_dimensions[r].height = 90
    values = [s["no"], s["time"], s["topic"],
              s["content"], s["practice"], s["hw"]]
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill      = fill(row_fills[i])
        cell.border    = border()
        cell.alignment = Alignment(wrap_text=True, vertical="top",
                                   horizontal="center" if c <= 3 else "left")
        if c == 1:
            cell.font = font(bold=True, size=12)
        elif c == 3:
            cell.font = font(bold=True, size=11)
        else:
            cell.font = font(size=10)

# ── Sheet 2: 作業清單 ───────────────────────────────────────
ws2 = wb.create_sheet("作業清單")
for col, w in zip(["A","B","C","D"], [8, 26, 54, 16]):
    ws2.column_dimensions[col].width = w

ws2.merge_cells("A1:D1")
h2 = ws2["A1"]
h2.value     = "⛏️  Minecraft Python 12堂作業清單"
h2.font      = Font(bold=True, size=14, color=WHITE, name="Calibri")
h2.fill      = fill(BROWN)
h2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

for c, h in enumerate(["堂次","作業名稱","作業說明","繳交期限"], 1):
    cell = ws2.cell(row=2, column=c, value=h)
    cell.font      = font(bold=True, color=WHITE)
    cell.fill      = fill("795548")
    cell.border    = border()
    cell.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[2].height = 22

hw_list = [
    ("第1堂",  "Minecraft 自我介紹",         "用 print() 印出角色卡（名稱、愛好、武器、目標），至少 10 行",                   "下次上課前"),
    ("第2堂",  "角色卡 v1（變數版）",         "建立 6 個不同型別的變數，模擬受傷（-7）、回血（+3）、升級（level+1）",           "下次上課前"),
    ("第3堂",  "附魔武器產生器",              "建立武器+附魔名稱+等級變數，用字串方法輸出格式化武器卡",                         "下次上課前"),
    ("第4堂",  "角色卡 v2（互動版）",         "用 input() 輸入角色資料，計算攻擊力 = 等級×3，用 f-string 印出角色卡",           "下次上課前"),
    ("第5堂",  "Minecraft 生存顧問",          "輸入血量/食物值/有無武器，根據不同組合給出對應建議",                            "下次上課前"),
    ("第6堂",  "Minecraft 合成台程式",        "輸入木頭/石頭/鐵錠/鑽石數量，判斷可合成哪些武器工具",                           "下次上課前"),
    ("第7堂",  "自動礦物統計機",              "for 迴圈跑 1~N 層，統計鑽石層/鐵礦層/岩漿層數量，用 continue 跳過岩漿層",        "下次上課前"),
    ("第8堂",  "Minecraft 生存模擬器",        "while 選單：挖礦/狩獵/吃東西，管理血量/食物/金幣，達標或歸零結束",               "下次上課前"),
    ("第9堂",  "完整背包管理系統",            "while 選單：新增/刪除/查詢/排序，背包上限 9 格",                                "下次上課前"),
    ("第10堂", "Minecraft RPG 工具包",        "寫 5 個函式：show_status/calc_damage/is_alive/can_craft/level_up",           "下次上課前"),
    ("第11堂", "Minecraft 地下城探索",        "3 個房間、每間有怪物+寶箱，戰鬥或逃跑，血量歸零遊戲結束",                       "下次上課前"),
    ("第12堂", "⭐ 期末專案：文字冒險遊戲",   "至少 3 個場景、怪物戰鬥、背包系統、勝利/失敗結局，綜合所有語法",                  "一週後"),
]

for i, (no, name, desc, dl) in enumerate(hw_list):
    r = i + 3
    ws2.row_dimensions[r].height = 52
    fc = GREEN_LIGHT if i % 2 == 0 else STONE_LIGHT
    for c, val in enumerate([no, name, desc, dl], 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.fill      = fill(fc)
        cell.border    = border()
        cell.alignment = Alignment(wrap_text=True, vertical="top",
                                   horizontal="center" if c in (1,2,4) else "left")
        cell.font = font(bold=(c==2), size=10)

# ── Sheet 3: 混合練習題 ─────────────────────────────────────
ws3 = wb.create_sheet("混合練習題")
for col, w in zip(["A","B","C","D"], [6, 18, 54, 34]):
    ws3.column_dimensions[col].width = w

ws3.merge_cells("A1:D1")
h3 = ws3["A1"]
h3.value     = "⚔️  Minecraft Python 混合練習題（for + if + while + function）"
h3.font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
h3.fill      = fill(GREEN_DARK)
h3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

for c, h in enumerate(["題號","難度","題目","提示"], 1):
    cell = ws3.cell(row=2, column=c, value=h)
    cell.font      = font(bold=True, color=WHITE)
    cell.fill      = fill(GREEN_MED)
    cell.border    = border()
    cell.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[2].height = 22

questions = [
    ("Q1","⭐ 基礎",
     "用 for 迴圈印出 1~10 層挖礦進度\n每層印：「第X層：開始挖掘...」",
     "for i in range(1, 11):\n    print(f\"第{i}層...\")"),
    ("Q2","⭐ 基礎",
     "輸入血量，if/elif/else 判斷狀態\n16-20=健康 / 8-15=一般 / 1-7=危險 / 0=死亡",
     "int(input()) + if/elif/else 四層"),
    ("Q3","⭐⭐ 普通",
     "while 猜數字遊戲（答案=64，代表鑽石層）\n猜太高/低給提示，猜中顯示共猜幾次",
     "while True + 計數器 + if/elif/else"),
    ("Q4","⭐⭐ 普通",
     "bag = ['木頭','石頭','鑽石','鐵礦','鑽石']\n用 for + if 計算背包裡有幾個鑽石",
     "for item in bag: if item == '鑽石': count+=1"),
    ("Q5","⭐⭐ 普通",
     "def analyze(y)：回傳該層的礦物\n  Y1-5→岩漿, Y6-12→鑽石, Y13-20→鐵礦, 其他→石頭\n主程式：for y in range(1,21) 印出每層",
     "def + for + if/elif"),
    ("Q6","⭐⭐⭐ 挑戰",
     "商店系統：\nbag=['木頭x10'] 金幣=30\nwhile 選單：木劍10金/石劍20金/鑽石劍50金\n金幣不夠警告，買完更新金幣",
     "while True + if/elif + int()"),
    ("Q7","⭐⭐⭐ 挑戰",
     "def fight(player_hp, enemy_hp, p_atk, e_atk)\n  while 兩者 hp > 0：輪流攻擊\n  回傳：'玩家勝' 或 '敵方勝'\n用三組不同數值呼叫",
     "def + while + if + return"),
    ("Q8","⭐⭐⭐⭐ 進階",
     "綜合：挖礦模擬器\n輸入要挖的層數 N\nfor 跑每層，if 判斷礦物（用亂數：import random）\n鑽石5%/金礦10%/鐵礦30%/煤礦40%/石頭15%\n最後用 dict 或多個變數統計各礦物數量",
     "import random + random.random() + for + if/elif + 計數器"),
]

q_fills = [STONE_LIGHT, STONE_LIGHT, GREEN_LIGHT, GREEN_LIGHT,
           GREEN_LIGHT, GOLD_LIGHT, GOLD_LIGHT, "FFCCBC"]
for i, (no, diff, q, hint) in enumerate(questions):
    r = i + 3
    ws3.row_dimensions[r].height = 70
    for c, val in enumerate([no, diff, q, hint], 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.fill      = fill(q_fills[i])
        cell.border    = border()
        cell.alignment = Alignment(wrap_text=True, vertical="top",
                                   horizontal="center" if c <= 2 else "left")
        cell.font = font(bold=(c==1), size=10)

# ── Sheet 4: 遊戲解鎖對照表 ────────────────────────────────
ws4 = wb.create_sheet("遊戲解鎖對照表")

ws4.merge_cells("A1:C1")
h4 = ws4["A1"]
h4.value     = "🎮  minecraft_quiz.py 解鎖代碼對照表（老師專用）"
h4.font      = Font(bold=True, size=14, color=WHITE, name="Calibri")
h4.fill      = fill(BROWN)
h4.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 30

for c, (h, w) in enumerate(zip(["上完這幾堂後","傳給學生的解鎖碼","解鎖的遊戲關卡"],
                                [16, 18, 36]), 1):
    cell = ws4.cell(row=2, column=c, value=h)
    cell.font      = font(bold=True, color=WHITE)
    cell.fill      = fill("4CAF50")
    cell.border    = border()
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws4.column_dimensions[get_column_letter(c)].width = w
ws4.row_dimensions[2].height = 24

unlock_data = [
    ("第1堂（預設）", "（不需要代碼，自動解鎖）",  "🟫 第一關：變數與資料型別"),
    ("第4堂後",       "STONE04",                   "💬 第二關：輸入輸出 & 字串"),
    ("第6堂後",       "IRON06",                    "⚔️  第三關：條件判斷"),
    ("第7堂後",       "GOLD07",                    "🔄  第四關：for 迴圈"),
    ("第8堂後",       "LAPIS08",                   "🔁  第五關：while 迴圈"),
    ("第9堂後",       "ENDER09",                   "🎒  第六關：串列 List"),
    ("第10堂後",      "NETHER10",                  "🔧  第七關：函式 Function"),
    ("第12堂後",      "DRAGON12",                  "🏆  BOSS關：綜合練習"),
]

for i, (after, code, level) in enumerate(unlock_data):
    r = i + 3
    ws4.row_dimensions[r].height = 36
    fc = GREEN_LIGHT if i % 2 == 0 else STONE_LIGHT
    for c, val in enumerate([after, code, level], 1):
        cell = ws4.cell(row=r, column=c, value=val)
        cell.fill      = fill(fc)
        cell.border    = border()
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.font = font(bold=(c in (1, 2)), size=10)

# 說明文字
r = len(unlock_data) + 4
ws4.merge_cells(f"A{r}:C{r}")
note = ws4.cell(row=r, column=1,
    value="操作方式：上完對應的課程後，把代碼傳給學生（LINE / 訊息均可）。學生在遊戲主選單按 [U] 輸入代碼，永久解鎖，代碼不分大小寫。")
note.font = font(size=10, color="555555")
note.alignment = Alignment(wrap_text=True)
ws4.row_dimensions[r].height = 36

out = "/Users/danny.chao/Desktop/python_class/teacher/Python課綱_Minecraft版.xlsx"
wb.save(out)
print(f"✅ 已儲存：{out}")
